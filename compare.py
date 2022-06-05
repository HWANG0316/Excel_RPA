def isNan(num):
    return num != num

def compare_excel(URL,old_xlsx, new_xlsx, column_name,count):
    import pandas as pd    
    import datetime
    import shutil
    import win32com.client
    from openpyxl import load_workbook
    global number_count_list

    update_number =  0
    delete_number = 0
    change_number = 0
    
    shutil.copy(new_xlsx, excel_file_name[count] + '_비교결과' + '.xlsx')      # 엑셀 파일 복사
    
    excel = win32com.client.Dispatch("Excel.Application")   # 엑셀 색칠하기 위한 객체 생성
    excel.Visible = False
    wb = excel.Workbooks.Open(URL + '/'+ excel_file_name[count] + '_비교결과' + '.xlsx')
    ws = wb.ActiveSheet

    
   
    df_old = pd.read_excel(old_xlsx,header = 3)     # 변경 전 데이터 읽어오기
    df_new = pd.read_excel(new_xlsx, header = 3)    # 변경 후 데이터 읽어오기 


    check = 1
    if len(df_old.values) != 0:
        for i in range(10):
            for j in range(10):
                if df_old.values[i][4] == df_new.values[j][4]:  # 여기부터 디버깅
                    check = check + 1
                    break

            
    if check < 5 and len(df_old.values) != 0 :
        
        df_old_list = []
        df_old_list_para3 = []
        one_list = []
        two_list = []
        three_list = []
        four_list = []

        one_list.extend(list(df_old["건물명"]))
        two_list.extend(list(df_old["실명칭"]))
        three_list.extend(list(df_old["층구분"]))
        four_list.extend(list(df_old["층수"]))

        for i in range(len(df_old)):
            df_old_list.append([one_list[i],two_list[i],three_list[i],four_list[i]])
            df_old_list_para3.append([one_list[i],two_list[i],three_list[i]])


        df_new_list = []
        df_new_list_para3 = []
        one_list = []
        two_list = []
        three_list = []
        four_list = []
        name_list = []
        one_list.extend(list(df_new["건물명"]))
        two_list.extend(list(df_new["실명칭"]))
        three_list.extend(list(df_new["층구분"]))
        four_list.extend(list(df_new["층수"]))

        for i in range(len(df_new)):
            df_new_list.append([one_list[i],two_list[i],three_list[i],four_list[i]])
            df_new_list_para3.append([one_list[i],two_list[i],three_list[i]])
            name_list.append(two_list[i])


        df_dropped = []
        df_added = []

        for i in range(len(df_old_list)):
                if df_old_list[i] in df_new_list:
                    pass
                else:
                    df_dropped.append(df_old_list[i])

        for i in range(len(df_new_list)):
            if df_new_list[i] in df_old_list:
                pass
            else:
                if df_new_list_para3[i] in df_old_list_para3: 
                    temp = df_old_list_para3.index(df_new_list_para3[i])

                    if df_new_list[i][3] != df_old_list[temp][3]:
                        if name_list.count(df_new_list[i][1]) >1: 
                            df_added.append(df_new_list[i])
                    else:
                        df_added.append(df_new_list[i])
                else:
                    df_added.append(df_new_list[i])            
                

        df_changed_old_list = df_old.values.tolist()
        df_changed_new_list = df_new.values.tolist()


        visited = [0 for i in range(len(df_changed_new_list))]

        for i in range(len(df_changed_old_list)):
            for j in range(len(df_changed_new_list)):
                if  df_changed_old_list[i][3] == df_changed_new_list[j][3] and visited[j] != 1:
                    visited[j] = 1
                    checked = 0
                    for k in range(4,56):
                        if df_changed_old_list[i][k] != df_changed_new_list[j][k] and (isNan(df_changed_old_list[i][k]) == False or isNan(df_changed_new_list[j][k]) == False): 
                            ws.Cells(j+5,k+1).Interior.ColorIndex = 27
                            
                            if checked != 1:
                                ws.Cells(j+5,1).Value = "변경"
                                change_number = change_number + 1
                                checked= 1
                    break
                else:
                    continue
        #삭제된 데이터



        #value_list = ws.Range("D5:D300").Value                      ### 엑셀 파일로부터 현재 존재하는 실코드명 추출해 리스트로 변형하는 부분
        #value_list = list(value_list)
        
        one_list = []
        two_list = []
        three_list = []
        four_list = []

        result_value_list = []
        for i in range(5,len(df_new_list)+5):
            result_value_list.append([ws.Cells(i,3).Value,ws.Cells(i,4).Value,ws.Cells(i,6).Value,int(ws.Cells(i,7).Value)])

        
        
        next_position = len(df_new_list) + 5     # 나중에 dropped 데이터 추가하기 위해서 위치를 저장함.
        
        for i in range(len(df_added)):                ##추가된 데이터가 존재한다면 결과 엑셀에서 추가된 데이터 색칠하기
            for j in range(len(result_value_list)):
                if df_added[i] == result_value_list[j] :
                    range_str = "A" +str(j+5) + ":BD" + str(j+5)
                    ws.Range(range_str).Interior.ColorIndex = 24
                    ws.Cells(j+5,1).Value = "추가"
                    update_number = update_number + 1
                    break
                
        wb.Save()
        excel.Quit()      #엑셀 닫기 
        
        
        load_wb = load_workbook(filename = excel_file_name[count] + '_비교결과' + '.xlsx',data_only = True)    # dropped data append위해 다시 파일을 연다.
        load_ws = load_wb.active
        ####### 여기부터 하면 됨.!!!!!!


        for i in range(len(df_dropped)):
            for j in range(len(df_old.values)):
                if df_dropped[i][1] == (df_old.values[j].tolist())[3]:
                    input_data = df_old.values[j].tolist()
                    load_ws.append(input_data)
        
        load_wb.save(excel_file_name[count] + '_비교결과' + '.xlsx')    
        
        excel.Visible = True
        wb = excel.Workbooks.Open(URL + '/' + excel_file_name[count] + '_비교결과' + '.xlsx')
        ws = wb.ActiveSheet
        
        
        for i in range(next_position,next_position + len(df_dropped)):
            range_str = "A" +str(i) + ":BD" + str(i)
            ws.Range(range_str).Interior.ColorIndex = 22
            ws.Cells(i,1).Value = "삭제"
            delete_number = delete_number + 1
        
        number_count_list.append([update_number,delete_number,change_number])

        wb.Save()
        excel.Quit()      #엑셀 닫기 
  
    else:
        df_old_list = []
        df_old_list_para3 = []
        one_list = []
        two_list = []
        three_list = []
        four_list = []
        five_list = []

        one_list.extend(list(df_old["건물명"]))
        #two_list.extend(list(df_old["실명칭"]))
        five_list.extend(list(df_old["실코드"]))
        three_list.extend(list(df_old["층구분"]))
        four_list.extend(list(df_old["층수"]))

        for i in range(len(df_old)):
            df_old_list.append([one_list[i],five_list[i],three_list[i],four_list[i]])
            df_old_list_para3.append([one_list[i],three_list[i]])


        df_new_list = []
        df_new_list_para3 = []
        one_list = []
        two_list = []
        three_list = []
        four_list = []
        five_list = []
        name_list = []

        one_list.extend(list(df_new["건물명"]))
        #two_list.extend(list(df_new["실명칭"]))
        five_list.extend(list(df_new["실코드"]))
        three_list.extend(list(df_new["층구분"]))
        four_list.extend(list(df_new["층수"]))

        for i in range(len(df_new)):
            df_new_list.append([one_list[i],five_list[i],three_list[i],four_list[i]])
            df_new_list_para3.append([one_list[i],three_list[i]])
            #name_list.append(two_list[i])


        df_dropped = []
        df_added = []

        for i in range(len(df_old_list)):
                if df_old_list[i] in df_new_list:
                    pass
                else:
                    df_dropped.append(df_old_list[i])

        for i in range(len(df_new_list)):
            if df_new_list[i] in df_old_list:
                pass
            else:
                df_added.append(df_new_list[i])            
                

        df_changed_old_list = df_old.values.tolist()
        df_changed_new_list = df_new.values.tolist()


        visited = [0 for i in range(len(df_changed_new_list))]

        for i in range(len(df_changed_old_list)):
            for j in range(len(df_changed_new_list)):
                if  df_changed_old_list[i][2] == df_changed_new_list[j][2] and  df_changed_old_list[i][4] == df_changed_new_list[j][4]and df_changed_old_list[i][5] == df_changed_new_list[j][5] and visited[j] != 1:
                    visited[j] = 1
                    checked = 0
                    for k in range(3,56):
                        if df_changed_old_list[i][k] != df_changed_new_list[j][k] and (isNan(df_changed_old_list[i][k]) == False or isNan(df_changed_new_list[j][k]) == False): 
                            ws.Cells(j+5,k+1).Interior.ColorIndex = 27

                            if checked != 1 :
                                ws.Cells(j+5,1).Value = "변경"
                                change_number = change_number + 1
                                checked = 1

                    break
                else:
                    continue
        #삭제된 데이터



        #value_list = ws.Range("D5:D300").Value                      ### 엑셀 파일로부터 현재 존재하는 실코드명 추출해 리스트로 변형하는 부분
        #value_list = list(value_list)
        
        one_list = []
        two_list = []
        three_list = []
        four_list = []

        result_value_list = []
        for i in range(5,len(df_new_list)+5):
            result_value_list.append([ws.Cells(i,3).Value,ws.Cells(i,5).Value,ws.Cells(i,6).Value,int(ws.Cells(i,7).Value)])

        
        
        next_position = len(df_new_list) + 5     # 나중에 dropped 데이터 추가하기 위해서 위치를 저장함.
        
        for i in range(len(df_added)):                ##추가된 데이터가 존재한다면 결과 엑셀에서 추가된 데이터 색칠하기
            for j in range(len(result_value_list)):
                if df_added[i] == result_value_list[j] :
                    range_str = "A" +str(j+5) + ":BD" + str(j+5)
                    ws.Range(range_str).Interior.ColorIndex = 24
                    ws.Cells(j+5,1).Value = "추가"
                    update_number = update_number + 1
                    break
                
        wb.Save()
        excel.Quit()      #엑셀 닫기 
        
        
        load_wb = load_workbook(filename = excel_file_name[count] + '_비교결과' + '.xlsx',data_only = True)    # dropped data append위해 다시 파일을 연다.
        load_ws = load_wb.active
        ####### 여기부터 하면 됨.!!!!!!


        for i in range(len(df_dropped)):
            for j in range(len(df_old.values)):
                if df_dropped[i][0] == (df_old.values[j].tolist())[2] and  df_dropped[i][1] == (df_old.values[j].tolist())[4] and df_dropped[i][2] == (df_old.values[j].tolist())[5]and df_dropped[i][3] == (df_old.values[j].tolist())[6]:
                    input_data = df_old.values[j].tolist()
                    load_ws.append(input_data)
        
        load_wb.save(excel_file_name[count] + '_비교결과' + '.xlsx')
        
        excel.Visible = True
        wb = excel.Workbooks.Open(URL + '/' + excel_file_name[count] + '_비교결과' + '.xlsx')
        ws = wb.ActiveSheet
        
        
        for i in range(next_position,next_position + len(df_dropped)):
            range_str = "A" +str(i) + ":BD" + str(i)
            ws.Range(range_str).Interior.ColorIndex = 22
            ws.Cells(i,1).Value = "삭제"
            delete_number = delete_number + 1
            

        number_count_list.append([update_number,delete_number,change_number])

        wb.Save()
        excel.Quit()      #엑셀 닫기 


import os
from openpyxl import Workbook


global number_count_list

number_count_list = []


print(os.getcwd())
URL = os.getcwd()
path = "./"
file_list = os.listdir(path)
before_file = []
after_file = []
status_file = []
school_name = []

for i in file_list:                    ## 전후 파일 나누는
    cut_number = i.find('_')
    cut_number = i.find('_', cut_number + 1)

    if "전" in i[cut_number:]:
        before_file.append(i)
    elif "후" in i[cut_number:]:
        after_file.append(i)

real_before_file = []
real_after_file = []

excel_file_name = [0 for i in range(len(before_file))]

for i in range(len(before_file)):
    if 'xlsx' in before_file[i] or 'xls' in before_file[i]:
        real_before_file.append(before_file[i])

for i in range(len(after_file)):
    if 'xlsx' in after_file[i] or 'xls' in after_file[i]:
        real_after_file.append(after_file[i])

real_before_file.sort()
real_after_file.sort()

for i in range(len(real_before_file)):
    cut_number = real_before_file[i].find('_',2)
    cut_number = real_before_file[i].find('_',cut_number + 1)
    excel_file_name[i] = real_before_file[i][3:cut_number]
    school_name.append(excel_file_name[i])

count = 0


for i in range(len(real_before_file)):
    compare_excel(URL,real_before_file[i],real_after_file[i],"실명칭",count)
    count = count + 1  



for i in range(len(number_count_list)):
    print("학교 이름 : ",school_name[i],  number_count_list[i])

excel_add = []

for i in range(len(number_count_list)):
    excel_add.append([school_name[i], number_count_list[i][0],number_count_list[i][1],number_count_list[i][2]])
write_wb = Workbook()

write_ws = write_wb.active
write_ws.append(["학교명","추가된 개수","삭제된 개수","변경된 개수"])
for i in range(len(number_count_list)):
    write_ws.append(excel_add[i])


write_wb.save(URL+'/Summary.xlsx')
